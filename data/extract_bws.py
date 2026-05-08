#!/usr/bin/env python3
"""Extract exercise library + program structure from BWS workbooks.

The hidden KX/KY columns (310/311) on each "last sheet" hold the master video lookup
(~120 exercises). The visible sheets hold program structure: column A for main lifts,
column C for superset partners (with A1/A2/B1/B2 labels in column B).
"""
import json
import re
from pathlib import Path
from openpyxl import load_workbook

DOWNLOADS = Path("/Users/carriermac/Downloads")
OUT_DIR = Path("/Users/carriermac/Downloads/fit-tracker/data")

FILES = {
    "bws_3day_full_body":      DOWNLOADS / "(V1-3 day) Program Workout Tracker.xlsx",
    "bws_5day_split_v1":       DOWNLOADS / "Intermediate (V1-5 day) Program Workout Tracker.xlsx",
    "bws_4day_upper_lower":    DOWNLOADS / "Copy of Intermediate (V1-4 day) Workout Tracker.xlsx",
    "bws_5day_split_v3":       DOWNLOADS / "Copy of Intermediate (V3-5 day) Program Workout Tracker.xlsx",
    "bws_5day_30min":          DOWNLOADS / "Copy of Intermediate (30min-5 day) Program Workout Tracker.xlsx",
    "bws_4day_30min":          DOWNLOADS / "Copy of Intermediate (30min-4 day) Program Workout Tracker.xlsx",
    "bws_ab_routine":          DOWNLOADS / "Copy of Intermediate Ab Routine Tracker.xlsx",
    "ryhan_5day_personal":     DOWNLOADS / "Ryhan 5 Day Split.xlsx",
    "ryhan_4day_personal":     DOWNLOADS / "Ryhan 4 day plan.xlsx",
    "ryhan_3day_personal":     DOWNLOADS / "3 Day split.xlsx",
}

def extract_video_lookup(wb):
    """Find KX/KY video lookup table (cols 310/311) on whichever sheet has it."""
    table = {}
    for ws in wb.worksheets:
        for r in range(1, 1100):
            name = ws.cell(row=r, column=310).value
            url = ws.cell(row=r, column=311).value
            if isinstance(name, str) and isinstance(url, str) and 'http' in url:
                # strip trailing asterisk markers
                clean_name = name.rstrip('*').strip()
                if clean_name and clean_name not in table:
                    table[clean_name] = url.strip()
        if table:
            break  # found it on this sheet, no need to scan others
    return table

def extract_program_days(wb, video_lookup):
    """For each visible sheet, parse exercise blocks (col A main + col C superset)."""
    days = []
    for ws in wb.worksheets:
        exercises = []
        current_superset = None
        for r in range(1, min(ws.max_row + 1, 200)):
            a = ws.cell(row=r, column=1).value
            b = ws.cell(row=r, column=2).value
            c = ws.cell(row=r, column=3).value

            # Detect superset header in column A
            if isinstance(a, str) and a.strip().upper().startswith("SUPERSET"):
                current_superset = a.strip().split()[-1]  # "A" or "B"
                # if column C also has an exercise on this same row, capture it
                if isinstance(c, str) and len(c.strip()) > 2:
                    name = c.strip()
                    sets = collect_set_targets(ws, r, col=2)  # set rows are in col B
                    exercises.append(make_ex(name, sets, current_superset, b, video_lookup))
                continue

            # Main exercise row in column A (skip headers/sets)
            if isinstance(a, str):
                s = a.strip()
                if s in ('Date', 'WORKOUTS', '') or s.startswith('Set ') or s.lower() == 'see tutorial video':
                    continue
                # is this a real exercise? check next 1-3 rows for "Set N:"
                sets = collect_set_targets(ws, r, col=1)
                if sets:
                    current_superset = None  # main lift breaks superset
                    exercises.append(make_ex(s, sets, None, None, video_lookup))
                continue

            # Superset partner exercise in column C (col B has A1/A2/B1/B2)
            if isinstance(c, str) and len(c.strip()) > 2 and isinstance(b, str) and re.match(r'^[AB][12]$', b.strip()):
                name = c.strip()
                if name.lower() == 'see tutorial video': continue
                sets = collect_set_targets(ws, r, col=2)  # set rows in col B
                ss = current_superset or b.strip()[0]
                exercises.append(make_ex(name, sets, ss, b.strip(), video_lookup))

        days.append({"day_name": ws.title, "exercises": exercises})
    return days

def collect_set_targets(ws, exercise_row, col):
    """Read the set N: X-Y rows that follow an exercise name."""
    targets = []
    for offset in range(1, 6):
        v = ws.cell(row=exercise_row + offset, column=col).value
        if not isinstance(v, str): continue
        m = re.match(r'^Set\s*\d+\s*:\s*(.+)$', v.strip(), re.IGNORECASE)
        if m:
            targets.append(m.group(1).strip())
        elif v.strip().lower() == 'see tutorial video':
            continue
        else:
            break
    return targets

def make_ex(name, set_targets, superset, label, video_lookup):
    clean_name = name.rstrip('*').strip()
    return {
        "name": clean_name,
        "set_targets": set_targets,
        "superset": superset,
        "label": label,
        "video_url": video_lookup.get(clean_name),
    }

def main():
    master_videos = {}     # canonical exercise -> url
    all_exercises = {}     # all exercise names ever seen (with merge of video URLs)
    programs = {}

    for prog_key, path in FILES.items():
        if not path.exists():
            print(f"[skip] {path}")
            continue
        print(f"\n[load] {path.name}")
        wb = load_workbook(path, data_only=True, read_only=False)
        videos = extract_video_lookup(wb)
        master_videos.update(videos)
        print(f"  master video lookup: {len(videos)} entries")
        days = extract_program_days(wb, videos)
        for d in days:
            print(f"  - {d['day_name']}: {len(d['exercises'])} exercises")
            for ex in d['exercises']:
                key = ex['name'].lower().strip()
                if key not in all_exercises:
                    all_exercises[key] = {
                        "name": ex['name'],
                        "video_url": ex.get('video_url'),
                        "sources": set(),
                    }
                else:
                    if ex.get('video_url') and not all_exercises[key].get('video_url'):
                        all_exercises[key]['video_url'] = ex['video_url']
                all_exercises[key]['sources'].add(prog_key)
        programs[prog_key] = {"days": days}

    # Merge master video lookup into exercise list
    for canonical_name, url in master_videos.items():
        key = canonical_name.lower().strip()
        if key not in all_exercises:
            all_exercises[key] = {
                "name": canonical_name,
                "video_url": url,
                "sources": ["bws_master_library"],
            }
        else:
            if not all_exercises[key].get('video_url'):
                all_exercises[key]['video_url'] = url
            all_exercises[key]['sources'].add("bws_master_library")

    # Finalize
    ex_list = sorted(
        [{"name": v['name'], "video_url": v.get('video_url'), "sources": sorted(v['sources'])}
         for v in all_exercises.values()],
        key=lambda x: x['name'].lower()
    )

    out = {
        "exercises": ex_list,
        "programs": programs,
        "stats": {
            "total_unique_exercises": len(ex_list),
            "with_video": sum(1 for e in ex_list if e['video_url']),
            "without_video": sum(1 for e in ex_list if not e['video_url']),
            "master_lookup_size": len(master_videos),
        },
    }
    out_path = OUT_DIR / "bws_exercise_db.json"
    out_path.write_text(json.dumps(out, indent=2, default=str))
    print(f"\n[wrote] {out_path}")
    for k, v in out['stats'].items():
        print(f"  {k}: {v}")

if __name__ == "__main__":
    main()
